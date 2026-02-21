# ============================================
# FILE PROCESSOR - Extract text from files
# ============================================

import os
import json
import csv
import logging
from typing import Tuple
from dataclasses import dataclass

from constants import (
    MAX_FILE_SIZE_BYTES,
    MAX_EXTRACTED_TEXT_CHARS,
    MAX_PDF_PAGES,
    MAX_EXCEL_ROWS,
    MAX_CSV_ROWS,
    SUPPORTED_FILE_EXTENSIONS
)

logger = logging.getLogger(__name__)


@dataclass
class FileProcessResult:
    """Result from processing a file."""
    success: bool
    text: str
    file_type: str
    error: str = ""
    truncated: bool = False


def get_file_extension(file_path: str) -> str:
    """Get lowercase file extension."""
    return os.path.splitext(file_path)[1].lower()


def is_supported_file(file_path: str) -> bool:
    """Check if file type is supported."""
    ext = get_file_extension(file_path)
    return ext in SUPPORTED_FILE_EXTENSIONS


def check_file_size(file_path: str) -> Tuple[bool, str]:
    """
    Check if file size is within limits.

    Returns:
        (is_valid, error_message)
    """
    try:
        size = os.path.getsize(file_path)
        if size > MAX_FILE_SIZE_BYTES:
            size_mb = size / (1024 * 1024)
            max_mb = MAX_FILE_SIZE_BYTES / (1024 * 1024)
            return False, f"File too large ({size_mb:.1f}MB). Maximum is {max_mb:.0f}MB."
        return True, ""
    except Exception as e:
        return False, f"Cannot read file: {e}"


def truncate_text(text: str, max_chars: int = MAX_EXTRACTED_TEXT_CHARS) -> Tuple[str, bool]:
    """
    Truncate text if it exceeds limit.

    Returns:
        (text, was_truncated)
    """
    if len(text) <= max_chars:
        return text, False
    return text[:max_chars] + "\n\n...[Content truncated due to length]", True


def process_text_file(file_path: str) -> FileProcessResult:
    """Process plain text files (.txt, .md, .log, code files)."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()

        text, truncated = truncate_text(text)

        return FileProcessResult(
            success=True,
            text=text,
            file_type="text",
            truncated=truncated
        )
    except Exception as e:
        logger.error(f"Failed to read text file: {e}")
        return FileProcessResult(
            success=False,
            text="",
            file_type="text",
            error=str(e)
        )


def process_json_file(file_path: str) -> FileProcessResult:
    """Process JSON files."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Pretty print JSON
        text = json.dumps(data, indent=2, ensure_ascii=False)
        text, truncated = truncate_text(text)

        return FileProcessResult(
            success=True,
            text=text,
            file_type="json",
            truncated=truncated
        )
    except Exception as e:
        logger.error(f"Failed to read JSON file: {e}")
        return FileProcessResult(
            success=False,
            text="",
            file_type="json",
            error=str(e)
        )


def process_csv_file(file_path: str) -> FileProcessResult:
    """Process CSV files."""
    try:
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= MAX_CSV_ROWS:
                    rows.append(f"...[Truncated at {MAX_CSV_ROWS} rows]")
                    break
                rows.append(" | ".join(row))

        text = "\n".join(rows)
        text, truncated = truncate_text(text)

        return FileProcessResult(
            success=True,
            text=text,
            file_type="csv",
            truncated=truncated or len(rows) > MAX_CSV_ROWS
        )
    except Exception as e:
        logger.error(f"Failed to read CSV file: {e}")
        return FileProcessResult(
            success=False,
            text="",
            file_type="csv",
            error=str(e)
        )


def process_pdf_file(file_path: str) -> FileProcessResult:
    """Process PDF files using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        return FileProcessResult(
            success=False,
            text="",
            file_type="pdf",
            error="pdfplumber not installed. Run: pip install pdfplumber"
        )

    try:
        text_parts = []
        truncated = False

        with pdfplumber.open(file_path) as pdf:
            total_pages = len(pdf.pages)
            pages_to_read = min(total_pages, MAX_PDF_PAGES)

            for i in range(pages_to_read):
                page = pdf.pages[i]
                page_text = page.extract_text() or ""
                if page_text:
                    text_parts.append(f"--- Page {i + 1} ---\n{page_text}")

            if total_pages > MAX_PDF_PAGES:
                text_parts.append(f"\n...[Showing {MAX_PDF_PAGES} of {total_pages} pages]")
                truncated = True

        text = "\n\n".join(text_parts)
        text, char_truncated = truncate_text(text)

        return FileProcessResult(
            success=True,
            text=text,
            file_type="pdf",
            truncated=truncated or char_truncated
        )
    except Exception as e:
        logger.error(f"Failed to read PDF file: {e}")
        return FileProcessResult(
            success=False,
            text="",
            file_type="pdf",
            error=str(e)
        )


def process_docx_file(file_path: str) -> FileProcessResult:
    """Process Word documents using python-docx."""
    try:
        from docx import Document
    except ImportError:
        return FileProcessResult(
            success=False,
            text="",
            file_type="docx",
            error="python-docx not installed. Run: pip install python-docx"
        )

    try:
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)

        text, truncated = truncate_text(text)

        return FileProcessResult(
            success=True,
            text=text,
            file_type="docx",
            truncated=truncated
        )
    except Exception as e:
        logger.error(f"Failed to read DOCX file: {e}")
        return FileProcessResult(
            success=False,
            text="",
            file_type="docx",
            error=str(e)
        )


def process_excel_file(file_path: str) -> FileProcessResult:
    """Process Excel files using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return FileProcessResult(
            success=False,
            text="",
            file_type="excel",
            error="openpyxl not installed. Run: pip install openpyxl"
        )

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        truncated = False

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")

            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= MAX_EXCEL_ROWS:
                    text_parts.append(f"...[Truncated at {MAX_EXCEL_ROWS} rows]")
                    truncated = True
                    break

                # Convert row to string
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)
                    row_count += 1

        wb.close()
        text = "\n".join(text_parts)
        text, char_truncated = truncate_text(text)

        return FileProcessResult(
            success=True,
            text=text,
            file_type="excel",
            truncated=truncated or char_truncated
        )
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        return FileProcessResult(
            success=False,
            text="",
            file_type="excel",
            error=str(e)
        )


def process_file(file_path: str) -> FileProcessResult:
    """
    Main function to process any supported file.

    Args:
        file_path: Path to the file

    Returns:
        FileProcessResult with extracted text or error
    """
    # Check if file exists
    if not os.path.exists(file_path):
        return FileProcessResult(
            success=False,
            text="",
            file_type="unknown",
            error="File not found"
        )

    # Check file extension
    ext = get_file_extension(file_path)
    if not is_supported_file(file_path):
        return FileProcessResult(
            success=False,
            text="",
            file_type="unknown",
            error=f"Unsupported file type: {ext}"
        )

    # Check file size
    size_ok, size_error = check_file_size(file_path)
    if not size_ok:
        return FileProcessResult(
            success=False,
            text="",
            file_type=ext,
            error=size_error
        )

    # Process based on file type
    if ext == ".pdf":
        return process_pdf_file(file_path)
    elif ext == ".docx":
        return process_docx_file(file_path)
    elif ext in [".xlsx", ".xls"]:
        return process_excel_file(file_path)
    elif ext == ".csv":
        return process_csv_file(file_path)
    elif ext == ".json":
        return process_json_file(file_path)
    else:
        # Text files (txt, md, log, code files)
        return process_text_file(file_path)
